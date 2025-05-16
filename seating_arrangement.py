import os
import pandas as pd
import logging
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font

# Configure logging
logging.basicConfig(
    filename='errors.txt',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def read_input_files():
    """Read and validate all input sheets with proper error handling."""
    try:
        xls = pd.ExcelFile("input_data_tt.xlsx")
        required_sheets = ['in_timetable', 'in_course_roll_mapping', 'in_roll_name_mapping', 'in_room_capacity']
        if not all(sheet in xls.sheet_names for sheet in required_sheets):
            raise ValueError("Missing required sheets in input file")

        # Read and clean timetable
        timetable_df = pd.read_excel(xls, 'in_timetable')
        timetable_df['Date'] = pd.to_datetime(timetable_df['Date'], format='%Y-%m-%d')
        for col in ['Morning', 'Evening']:
            timetable_df[col] = timetable_df[col].apply(
                lambda x: [] if pd.isna(x) else [s.strip() for s in str(x).split(';') if s.strip()]
            ).apply(lambda lst: [s for s in lst if s.upper() != 'NO EXAM'])

        # Read and clean course-roll mapping
        course_roll_df = pd.read_excel(xls, 'in_course_roll_mapping')
        course_roll_df['course_code'] = course_roll_df['course_code'].astype(str).str.strip()
        course_roll_df['rollno'] = course_roll_df['rollno'].astype(str).str.strip()

        # Read and clean roll-name mapping (fix: column is "Roll" in Excel)
        roll_name_df = pd.read_excel(xls, 'in_roll_name_mapping')
        roll_name_df.columns = [c.strip() for c in roll_name_df.columns]
        if 'Roll' in roll_name_df.columns:
            roll_name_df.rename(columns={'Roll': 'rollno', 'Name': 'name'}, inplace=True)
        roll_name_df['rollno'] = roll_name_df['rollno'].astype(str).str.strip()
        roll_name_df['name'] = roll_name_df['name'].astype(str).str.strip()

        # Read and clean room capacity
        room_df = pd.read_excel(xls, 'in_room_capacity')
        room_df = room_df.iloc[:, :3]  # Select first 3 columns
        room_df.columns = ['Room No.', 'Exam Capacity', 'Block']
        room_df['Room No.'] = room_df['Room No.'].astype(str).str.strip()
        room_df['Block'] = room_df['Block'].astype(str).str.strip()

        return timetable_df, course_roll_df, roll_name_df, room_df

    except Exception as e:
        logging.error(f"Input file error: {str(e)}", exc_info=True)
        raise

def get_subjects_for_date_session(timetable_df, date_obj, session):
    """Get list of subjects scheduled for a specific date and session."""
    row = timetable_df[timetable_df['Date'] == date_obj]
    if row.empty:
        print(f"No timetable entry found for date: {date_obj}")
        return []
    subjects = row.iloc[0][session]
    return subjects

def check_clashes(subjects, course_roll_df):
    """Check if any student is enrolled in multiple subjects."""
    roll_sets = []
    for subj in subjects:
        rolls = set(course_roll_df[course_roll_df['course_code'] == subj]['rollno'])
        roll_sets.append(rolls)
    
    for i in range(len(roll_sets)):
        for j in range(i+1, len(roll_sets)):
            common = roll_sets[i] & roll_sets[j]
            if common:
                logging.error(f"Clash between {subjects[i]} and {subjects[j]}: {common}")
                print(f"Clash detected between {subjects[i]} and {subjects[j]} for rolls: {sorted(list(common))[:5]}...")
                return True
    print("No clashes detected.")
    return False

def allocate_subjects(subjects, course_roll_df, room_df, buffer, arrangement_type):
    """Optimized room allocation with building/floor priority."""
    # Prepare room data with floor extraction for proximity
    room_df = room_df.copy()
    room_df['floor'] = room_df['Room No.'].str.extract(r'(\d{2,})').astype(float)
    room_df['remaining'] = room_df['Exam Capacity'] - buffer
    
    if arrangement_type.lower() == 'sparse':
        room_df['remaining'] = room_df['remaining'] // 2
    room_df['remaining'] = room_df['remaining'].apply(lambda x: max(x, 0))
    
    # Sort rooms by block, floor, capacity (largest first)
    room_df = room_df.sort_values(['Block', 'floor', 'remaining'], 
                                 ascending=[True, True, False])

    allocation = defaultdict(list)
    unallocated = {}
    room_capacity = room_df.set_index('Room No.')['remaining'].to_dict()
    block_rooms = room_df.groupby('Block')['Room No.'].apply(list).to_dict()
    
    # Sort subjects by size (largest first)
    subject_sizes = {subj: len(course_roll_df[course_roll_df['course_code'] == subj]) 
                    for subj in subjects}
    sorted_subjects = sorted(subjects, key=lambda x: -subject_sizes[x])

    for subj in sorted_subjects:
        students = sorted(course_roll_df[course_roll_df['course_code'] == subj]['rollno'].tolist())
        remaining_students = students.copy()
        
        # Try to allocate within a single block first
        allocated = False
        for block in block_rooms:
            if not remaining_students:
                break
                
            for room in block_rooms[block]:
                if not remaining_students:
                    break
                    
                available = room_capacity.get(room, 0)
                if available <= 0:
                    continue
                    
                alloc = min(len(remaining_students), available)
                allocation[(subj, room)].extend(remaining_students[:alloc])
                room_capacity[room] -= alloc
                remaining_students = remaining_students[alloc:]
                
            if not remaining_students:
                allocated = True
                break

        # Fallback to any available capacity
        if not allocated and remaining_students:
            for room in room_df['Room No.']:
                if not remaining_students:
                    break
                    
                available = room_capacity.get(room, 0)
                if available <= 0:
                    continue
                    
                alloc = min(len(remaining_students), available)
                allocation[(subj, room)].extend(remaining_students[:alloc])
                room_capacity[room] -= alloc
                remaining_students = remaining_students[alloc:]

        if remaining_students:
            unallocated[subj] = remaining_students
            logging.warning(f"Couldn't allocate {len(remaining_students)} students for {subj}")
            print(f"Cannot allocate {len(remaining_students)} students for {subj} (excess students).")
    
    return allocation, unallocated, room_capacity

def format_room_excel(filepath, date_str, session, room, course, df):
    """Format room Excel file with headers and TA/Invigilator placeholders."""
    wb = Workbook()
    ws = wb.active
    
    # Add merged header row
    header_text = f"Exam Date: {date_str} | Session: {session} | Room: {room} | Course: {course}"
    ncols = len(df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=header_text)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, size=12)
    
    # Add student list
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Add TA/Invigilator placeholders
    start_row = ws.max_row + 2
    for i in range(1, 6):
        ws.cell(row=start_row + i - 1, column=1, value=f"TA {i}")
    for i in range(1, 6):
        ws.cell(row=start_row + 5 + i - 1, column=1, value=f"Invigilator {i}")
    
    wb.save(filepath)

def generate_outputs(date_str, day_str, session, allocation, roll_name_dict, room_df, room_capacity):
    """Generate all required output files and folders."""
    # Create directory structure
    base_dir = f"{date_str.replace('-', '')}_{day_str.replace(' ', '_')}"
    session_dir = os.path.join(base_dir, session.lower())
    os.makedirs(session_dir, exist_ok=True)
    
    # Prepare master data and per-room files
    master_rows = []
    for (subj, room), rolls in allocation.items():
        # Generate per-room Excel with formatting
        df = pd.DataFrame({
            'Roll Number': rolls,
            'Name': [roll_name_dict.get(r, "Unknown Name") for r in rolls]
        })
        filename = f"{subj}_{room}.xlsx"
        filepath = os.path.join(session_dir, filename)
        format_room_excel(filepath, date_str, session, room, subj, df)
        
        # Add to master data
        master_rows.append({
            'Date': date_str,
            'Day': day_str,
            'course_code': subj,
            'Room': room,
            'Allocated_students_count': len(rolls),
            'Roll_list (semicolon separated_)': ';'.join(rolls)
        })
    
    # Update master file
    master_file = 'op_overall_seating_arrangement.xlsx'
    if os.path.exists(master_file):
        existing = pd.read_excel(master_file)
        master_df = pd.concat([existing, pd.DataFrame(master_rows)], ignore_index=True)
    else:
        master_df = pd.DataFrame(master_rows)
    master_df.to_excel(master_file, index=False)
    
    # Generate seats left file
    seats_left = []
    for _, row in room_df.iterrows():
        room = row['Room No.']
        seats_left.append({
            'Room No.': room,
            'Exam Capacity': row['Exam Capacity'],
            'Block': row['Block'],
            'Alloted': int(row['Exam Capacity'] - room_capacity.get(room, row['Exam Capacity'])),
            'Vacant (B-C)': int(room_capacity.get(room, row['Exam Capacity']))
        })
    pd.DataFrame(seats_left).to_excel('op_seats_left.xlsx', index=False)

def test_allocation(date_str='2016-04-30', session='Morning', buffer=5, arrangement='sparse'):
    """Test allocation for a specific date and session."""
    try:
        # Load data
        timetable_df, course_roll_df, roll_name_df, room_df = read_input_files()
        roll_name_dict = roll_name_df.set_index('rollno')['name'].to_dict()
        
        date_obj = pd.to_datetime(date_str)
        print(f"Running test allocation for {date_str} {session} session...")
        
        # Get subjects and check clashes
        subjects = get_subjects_for_date_session(timetable_df, date_obj, session)
        print(f"Subjects found: {subjects}")
        
        if check_clashes(subjects, course_roll_df):
            print("Clash detected - cannot proceed.")
            return
        
        # Run allocation
        allocation, unallocated, remaining_cap = allocate_subjects(
            subjects, course_roll_df, room_df, buffer, arrangement
        )
        
        # Print results
        print("\n--- Allocation Results ---")
        for (subj, room), students in allocation.items():
            print(f"{subj} in {room}: {len(students)} students")
            
        print("\n--- Remaining Capacities ---")
        for room, cap in remaining_cap.items():
            print(f"{room}: {cap} seats left")
            
        if unallocated:
            print("\n--- Unallocated Students ---")
            for subj, students in unallocated.items():
                print(f"{subj}: {len(students)} students")
                
        # Generate outputs
        day_str = timetable_df[timetable_df['Date'] == date_obj].iloc[0]['Day']
        generate_outputs(date_str, day_str, session, allocation, roll_name_dict, room_df, remaining_cap)
        
    except Exception as e:
        logging.error(f"Error in test_allocation: {str(e)}", exc_info=True)
        print(f"Error: {str(e)}")

def main():
    """Main execution flow with user inputs."""
    try:
        # Load data
        timetable_df, course_roll_df, roll_name_df, room_df = read_input_files()
        roll_name_dict = roll_name_df.set_index('rollno')['name'].to_dict()
        
        print("Data loaded successfully:")
        print(f"- Timetable entries: {len(timetable_df)}")
        print(f"- Unique courses: {course_roll_df['course_code'].nunique()}")
        print(f"- Rooms: {len(room_df)}")
        
        # User inputs
        buffer = int(input("Enter buffer (seats to leave empty): "))
        arrangement_type = input("Enter allocation type (sparse/dense): ").lower()
        while arrangement_type not in ['sparse', 'dense']:
            arrangement_type = input("Invalid input! Enter 'sparse' or 'dense': ").lower()

        # Process each date and session
        for _, row in timetable_df.iterrows():
            date_str = row['Date'].strftime('%Y-%m-%d')
            day_str = row['Day']
            
            for session in ['Morning', 'Evening']:
                subjects = row[session]
                if not subjects:
                    continue
                
                print(f"\nProcessing {date_str} {session}...")
                
                # Check clashes
                if check_clashes(subjects, course_roll_df):
                    print("Clash detected. Skipping allocation for this slot.")
                    continue
                
                # Allocate rooms
                allocation, unallocated, remaining_cap = allocate_subjects(
                    subjects, course_roll_df, room_df, buffer, arrangement_type
                )
                
                # Generate outputs
                generate_outputs(date_str, day_str, session, allocation, roll_name_dict, room_df, remaining_cap)
                
                print(f"Processed {len(subjects)} subjects for {session} session.")
                if unallocated:
                    print(f"Warning: Couldn't allocate {sum(len(v) for v in unallocated.values())} students")

        print("\nAllocation completed successfully!")
        
    except Exception as e:
        logging.critical(f"Fatal error: {str(e)}", exc_info=True)
        print("Critical error occurred! Check errors.txt for details.")

if __name__ == "__main__":
    main()
